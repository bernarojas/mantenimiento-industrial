from django.http import HttpResponse
import datetime
from django.template import Template, Context
from django.template.loader import get_template
from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from predecirFotos.models import EQUIPOSCOLOSO, Sensores_excel, SENSORES
from datetime import datetime
import matplotlib.pyplot as plt
from django.utils import timezone
import dateutil.parser
from django.views.generic import TemplateView
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from .filters import OrderFilter
import numpy as np
from numpy import array
from pytz.exceptions import AmbiguousTimeError
import collections
import os
import matplotlib.pyplot as plt
from scipy.stats import pearsonr
from scipy.stats import spearmanr

class InicioExcel(TemplateView):
    template_name = 'DatosExcel.html'

def CargaWebExcel(request):
    #SENSORES.objects.all().delete()
    #Sensores_excel.objects.all().delete()
    if request.POST["equipoSelect"] == '4':
        EQUIPOSCOLOSO.objects.get_or_create(Nombre_equipo='Bomba 501', Descripcion='ESP3 - Baja Densidad',
                                Tipo_equipo='PTO.BOMBA')
        print("Imprimiendo SELECT:")
        print(request.POST["equipoSelect"])
        Sensores_excel.objects.all().delete()
        Excel = request.FILES["datosExcel"]
        wb = load_workbook(Excel)
        sheet_obj = wb.active
        numeroFilas = sheet_obj.max_row

        for e in range(2, numeroFilas+1):
                        #GuardarSensoresExcel=Sensores_excel( Fecha = sheet_obj.cell(row = e, column = 1).value, 
                        #                                Funcionando = sheet_obj.cell(row = e, column = 2).value,
                        #                                Presion_agua_sello = sheet_obj.cell(row = e, column = 3).value,
                        #                                Flujo_agua_sello = sheet_obj.cell(row = e, column = 4).value,
                        #                                Velocidad = sheet_obj.cell(row = e, column = 5).value,
                        #                                Flujo_transmisor = sheet_obj.cell(row = e, column = 6).value,
                        #                                Densimetro_nuclear = sheet_obj.cell(row = e, column = 7).value,
                        #                               Temperatura = 0, Aceleracion = 0, Velocidad_vibracion = 0, 
                        #                                Equipo1_id= 1) 
                        GuardarSensoresExcel=Sensores_excel.objects.get_or_create( Fecha = sheet_obj.cell(row = e, column = 1).value, 
                                                        Funcionando = sheet_obj.cell(row = e, column = 2).value,
                                                        Presion_agua_sello = sheet_obj.cell(row = e, column = 3).value,
                                                        Flujo_agua_sello = sheet_obj.cell(row = e, column = 4).value,
                                                        Velocidad = sheet_obj.cell(row = e, column = 5).value,
                                                        Flujo_transmisor = sheet_obj.cell(row = e, column = 6).value,
                                                        Densimetro_nuclear = sheet_obj.cell(row = e, column = 7).value,
                                                        Temperatura = 0, Aceleracion = 0, Velocidad_vibracion = 0, 
                                                        Equipo1_id= 1)
                        #GuardarSensoresExcel.save()

    if request.POST["equipoSelect"] == '1':
        EQUIPOSCOLOSO.objects.get_or_create(Nombre_equipo='Bomba 501', Descripcion='ESP3 - Baja Densidad',
                                            Tipo_equipo='PTO.BOMBA')
        print("Imprimiendo SELECT:")
        print(request.POST["equipoSelect"])
        SENSORES.objects.all().delete()
        Excel = request.FILES["datosExcel"]
        wb = load_workbook(Excel)
        sheet_obj = wb.active
        numeroFilas = sheet_obj.max_row

        for e in range(2, numeroFilas+1):
                    GuardarSensoresExcel=SENSORES.objects.get_or_create( Fecha = sheet_obj.cell(row = e, column = 1).value, 
                                                        Presion_agua_sello = sheet_obj.cell(row = e, column = 3).value,
                                                        Flujo_agua_sello = sheet_obj.cell(row = e, column = 4).value,
                                                        Velocidad = sheet_obj.cell(row = e, column = 5).value,
                                                        Flujo_transmisor = sheet_obj.cell(row = e, column = 6).value,
                                                        Densimetro_nuclear = sheet_obj.cell(row = e, column = 7).value,
                                                        Temperatura = sheet_obj.cell(row = e, column = 8).value, 
                                                        Aceleracion = sheet_obj.cell(row = e, column = 9).value,
                                                        Velocidad_vibracion = sheet_obj.cell(row = e, column = 10).value, 
                                                        Equipo_id= 1, Ingreso = 'Manual')
                    #GuardarSensoresExcel.save()                     
    #return redirect('resultadoTesteo2.html')
    #return HttpResponseRedirect('resultadoTesteo2.html')
    return render(request, 'resultadoTesteo2.html')


class FiltrosExcel(TemplateView):
    template_name = 'filtros.html'

